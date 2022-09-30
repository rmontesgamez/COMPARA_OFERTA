#import pyodbc
import openpyxl as op
import datetime
from consulta import consulta
import volcador
import pandas as pd
#import dato
import os
#from openpyxl import Workbook

#from openpyxl.styles import Color, PatternFill, Font, Border
#from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting.rule import ColorScaleRule , CellIsRule, FormulaRule
#from openpyxl.formatting import Rule


def datos_piezas(tipo, numero_cliente, estado_pedido):
    fecha=datetime.datetime.today()
    #fecha -= datetime.timedelta(days=365)
    fecha_str = fecha.strftime('%m/%d/%Y')

    parametros_consulta1={'fecha_inicial':fecha_str, 'fecha_final': fecha_str, 'n_cliente': numero_cliente, 'estado_pedido': estado_pedido}
    cr=consulta()
    frase=cr.sql_query(tipo, parametros_consulta1)

    datos=cr.consulta_pandas(frase)
   
    return datos 

fecha=datetime.datetime.today()
fecha_str = fecha.strftime('%m/%d/%Y')

parametros_consulta1={}

nombre_archivo="RESUMEN.xlsx"   


# borrado de archivo anterior
ruta2="C:\\activa\\pruebas_and\\"+ nombre_archivo
if os.path.exists(ruta2):
    os.remove(ruta2)


# recuperación datos de archivo de texto
datos = []
with open ("C:\\activa\\datoconsulta.txt",'r') as archivo:
    lineas = archivo.readlines()
    
    for linea in lineas:
        datos.append(linea.strip('\n'))
          
print(datos) # solo para pruebas
archivo.close()

destino = datos[0] # correo desde el que se envia la solicitud

# tratamiento datos de archivo
if datos[1]=='Cliente':
    
    if datos[2] == 'AND&OR':
        num_cliente='03217'
    elif datos[2] == 'MOLDTECH':
        num_cliente = '02366'
    else:
        num_cliente = '00000'

    if num_cliente == '00000':
        pedidos_no_cerrados = []
        pedidos_plegado = []
    else:
        num_pedido = ''
        pedidos_no_cerrados =  datos_piezas('pedidos_n_cerrados', num_cliente, num_pedido)
        pedidos_plegado = datos_piezas('pedidos_plegado', num_cliente, num_pedido)
        
    datos_conjunto1 = pedidos_no_cerrados
    datos_conjunto2 = pedidos_plegado
        


elif datos[1] == 'Pedido':
    
    num_pedido = datos[2]
    num_cliente = ''
    est_ped_no_cerrados =  datos_piezas('estado_pedido', num_cliente , num_pedido)
    est_ped_pleg = datos_piezas('est_pedidos_pleg', num_cliente , num_pedido)
    

    datos_conjunto1 = est_ped_no_cerrados
    datos_conjunto2 = est_ped_pleg


elif datos[1] == 'Pieza':
    
    num_pedido = datos[2]
    num_cliente = ''
    est_pieza_no_cerrados =  datos_piezas('estado_pieza', num_cliente , num_pedido)
    est_pieza_pleg = datos_piezas('est_piezas_pleg', num_cliente , num_pedido)
    
    datos_conjunto1 = est_pieza_no_cerrados
    datos_conjunto2 = est_pieza_pleg




  #fecha.strftime('%Y%m%d')+ ".xlsm"
#nombre_cortadas='CORTADAS.xlsx'
#nombre_plegado='PLEGADO_NO_C.xlsx'
#nombre_no_cerrads = "NO_CERRADO.xlsx"


#ruta_cortadas="C:\\activa\\pruebas_and\\"+ nombre_cortadas
#ruta_plegado="C:\\activa\\pruebas_and\\"+ nombre_plegado
#ruta_no_cerrado = "C:\\activa\\pruebas_and\\"+ nombre_no_cerrads

if len(datos_conjunto1):
    if len(datos_conjunto2):
        datos_conjunto2['PDT_PLEG'] = datos_conjunto2.apply(lambda fila: (fila['CANT']-fila['CANTREALZDAS']), axis = 1)
        datos_conjunto2 = datos_conjunto2.drop(['CANT', 'CANTREALZDAS'], axis=1)

        datos_conjuntos3 = pd.merge(datos_conjunto1, datos_conjunto2, on =['CODCLIENTE','PEDIDO', 'CODPZPEDIDO', 'O_PED', 'TRATMTO', 'REF', 'REF_C'], how = 'outer')
    
    else:
        datos_conjuntos3 = datos_conjunto1
        datos_conjuntos3['PDT_PLEG'] = ''


    ordenado = datos_conjuntos3.sort_values(['PEDIDO', 'O_PED'])

    ordenado = ordenado.drop(['CODCLIENTE', 'CODPZPEDIDO'], axis=1)

    ordenado.rename(columns = {'REF': 'REF_PEDIDO', 'REF_C' : 'REF_CLIENTE', 'QP' : 'CANT_PEDIDA', 'PDTESP' : 'PDTE_PROG', 'PDTESC' : 'PDTE_COR'}, inplace=True)
    
       

    volcador.volcado_con_pandas(ordenado,'PIEZAS',ruta2, "b" )


    wb = op.load_workbook(ruta2)
    ws = wb.active


    longitud = len(ws['A'])
    print(longitud)

    lista_columna = ['H', 'I', 'J']
    for letra in lista_columna:
        campo = letra + '2:' + letra + str(longitud)

        redFill = PatternFill(start_color='EE1111',
                       end_color='EE1111',
                       fill_type='solid')
        greenfill = PatternFill(start_color='00F000',
                       end_color='00F000',
                       fill_type='solid' )
        ws.conditional_formatting.add(campo,
            FormulaRule(formula=[(letra +'2>0')], fill=redFill))
        ws.conditional_formatting.add(campo,
            FormulaRule(formula=[(letra + '2<=0')], fill=greenfill))

    wb.save(ruta2)
    volcador.correo(ruta2,'RESP. AUTOMATICA', 'RESUMEN.xlsx', destino)
else:
    ruta3="C:\\activa\\pruebas_and\\ERROR SOLICITUD.txt"

    volcador.correo(ruta3, 'SIN DATOS', 'ERROR.txt', destino)