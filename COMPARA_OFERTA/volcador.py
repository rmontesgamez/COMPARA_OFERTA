
import openpyxl as op
import smtplib
import imghdr
import pandas as pd


def ofertado(libro,listado, hoja):
    sheet_hoja = libro.create_sheet(hoja,0)

    print(libro.sheetnames)

    sheet_hoja.append(['CLIENTE', 'PRECIO TOT', 'ESTADO', 'OFERTA','TELÉFONO', 'FECHA'])

    sheet_hoja.column_dimensions['A'].width=50
    lista=['B','C','D', 'E', 'F']

    for letra in lista:
        sheet_hoja.column_dimensions[letra].width=15

    if len(listado)>1:
        print(len(listado))

        for row in listado:
            row2=list(row)
                        
            row2[5]=row2[5].strftime('%d/%m/%Y')

            if row2[2]==0:
               row2[2]="Por revisar"
            elif row2[2]==2:
               row2[2]="Enviado"
            elif row2[2]==3:
               row2[2]="Rechazado"
            elif row2[2]==1:
               row2[2]="Por enviar"
            elif row2[2]==4:
                row2[2]="Aceptado"
            

            sheet_hoja.append(row2)


def pedido(libro, listado):
    sheet_hoja = libro.create_sheet("PEDIDO",0)

    print(libro.sheetnames)


    sheet_hoja.append(['CLIENTE', 'PEDIDO', 'REFERENCIA'])

    sheet_oferta.column_dimensions['A'].width=50
    lista=['B','C','D', 'E', 'F']

    for letra in lista:
        sheet_oferta.column_dimensions[letra].width=20

    if len(listado)>1:
        print(len(listado))

        for row in listado:
            row2=list(row)
              
            sheet_hoja.append(row2)

def oferta_detallada(libro, listado, hoja):
    sheet_hoja = libro.create_sheet(hoja,0)

    print(libro.sheetnames)

    sheet_hoja.append(['CLIENTE', 'OFERTA', 'PRECIO TOT', 'PZO UNIT', 'CANT','REF', 'MATERIAL', 'TRATAMIENTO'])

    sheet_hoja.column_dimensions['A'].width=50
    sheet_hoja.column_dimensions['F'].width=50

    lista=['B','C','D', 'E', 'G', 'H']

    for letra in lista:
        sheet_hoja.column_dimensions[letra].width=15

    if len(listado)>1:
        print(len(listado))

        for row in listado:
            row2=list(row)
              
            sheet_hoja.append(row2)


def consulta_generica(libro,listado, hoja):
    sheet_hoja = libro.create_sheet(hoja,0)

    print(libro.sheetnames)

  
    if len(listado)>1:
        print(len(listado))

        for row in listado:
            row2=list(row)

            sheet_hoja.append(row2)


def consulta_clientes(libro,listado, hoja):
    sheet_hoja = libro.create_sheet(hoja,0)

    print(libro.sheetnames)
    sheet_hoja.append(['CLIENTE', 'CÓDIGO', 'DIRECCIÓN', 'NOMBRE', 'TELÉFONO'])

    lista=['A', 'B','C','D', 'E']

    for letra in lista:
        sheet_hoja.column_dimensions[letra].width=40
  
    if len(listado)>1:
        print(len(listado))

        for row in listado:
            row2=list(row)

            sheet_hoja.append(row2)



def correo(archivo, asunto, nombre_archivo, destinatario):
    from email.message import EmailMessage
    Sender_Email = "laserguadalquivir.rafa@gmail.com"
    Reciever_Email = destinatario

    #Reciever_Email = "gerencia@laserguadalquivir.com"

    Password ='Laser1234.'
    newMessage = EmailMessage()                         
    newMessage['Subject'] = asunto
    newMessage['From'] = Sender_Email                   
    newMessage['To'] = Reciever_Email                   
    newMessage.set_content('RESP. AUTOMATICA') 
    files = [archivo]
    for file in files:
        with open(file, 'rb') as f:
            file_data = f.read()
            file_name = f.name
        newMessage.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=nombre_archivo)
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(Sender_Email, Password)              
        smtp.send_message(newMessage)
        smtp.quit()


def volcado_con_pandas(listado, hoja, ruta, columna_celda):
    

    datos = pd.DataFrame(listado)

    datos.to_excel(ruta, sheet_name= hoja, index=False, header=True)



def volcado_2h_pandas(listado1, hoja1, listado2, hoja2, ruta, columna_celda):
    #sheet_hoja = libro.create_sheet(hoja,0)
    #libro.save(ruta)  
    #print(libro.sheetnames)
    writer = pd.ExcelWriter(ruta)

    datos1 = pd.DataFrame(listado1)

    datos1.to_excel(writer, sheet_name= hoja1, index=False, header=True)

    datos2 = pd.DataFrame(listado2)

    datos2.to_excel(writer, sheet_name= hoja2, index=False, header=True)

    writer.save()
    writer.close()

def volcado_reutiliza_excel(listado, hoja, ruta):
    datos = pd.DataFrame(listado)
    book = op.load_workbook("P:\\Personal Láser Guadalquivir\\RAFA\PLEGADO\\RESUMEN1.xlsm", keep_vba = True) # Load existing .xlsm file

    with pd.ExcelWriter(ruta, engine='openpyxl') as writer: # open a writer instance with the filename of the 
    
        writer.book = book # Hand over input workbook
        #writer.sheets = dict((ws.title, ws) for ws in book.worksheets) # Hand over worksheets
        writer.vba_archive = book.vba_archive # Hand over VBA information 
        datos.to_excel(writer, sheet_name = hoja,
                 header = True , index = False)



        writer.save()

    book = op.load_workbook("P:\\Personal Láser Guadalquivir\\RAFA\PLEGADO\\RESUMEN.xlsm", keep_vba = True)
    book.save(ruta)
    
    
        #writer.close()

