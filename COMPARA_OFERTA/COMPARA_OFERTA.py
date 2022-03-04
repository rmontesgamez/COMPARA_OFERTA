import pyodbc
import openpyxl as op
import datetime
from consulta import consulta
import volcador
import pandas as pd
import dato
import os
from openpyxl import Workbook
import plano
from pathlib import Path

from openpyxl.styles import Color, PatternFill, Font, Border
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule


def datos_piezas(tipo, numero_cliente, estado_pedido):
    fecha=datetime.datetime.today()
    #fecha -= datetime.timedelta(days=365)
    fecha_str = fecha.strftime('%m/%d/%Y')

    parametros_consulta1={'fecha_inicial':fecha_str, 'fecha_final': fecha_str, 'n_cliente': numero_cliente, 'estado_pedido': estado_pedido}
    cr=consulta()
    frase=cr.sql_query(tipo, parametros_consulta1)

    datos=cr.consulta_pandas(frase)
   
    return datos 

fecha=datetime.datetime.today()
fecha_str = fecha.strftime('%m/%d/%Y')

parametros_consulta1={}
# ruta donde se crea el archivo
nombre_archivo="COMPARA_OFERTA.xlsx"   


# borrado de archivo anterior
ruta2="C:\\activa\\"+ nombre_archivo
if os.path.exists(ruta2):
    os.remove(ruta2)

ruta_carpeta = "C:\\activa\\PKS\\"

if os.path.isdir(ruta_carpeta):
    directorio=Path(ruta_carpeta)

    for fichero in directorio.iterdir():
        if fichero.is_file():
            try:
                #print(fichero)
                os.remove(fichero)
            except:
                print("Error en fichero", fichero.name)




# comprobamos el número oferta
mayor_oferta = datos_piezas('max_oferta', '', '')
while True:
    try:
        oferta = int(input("INTRODUCIR NÚMERO OFERTA: "))
    except ValueError:
        print("OJO, DEBE SER UN NÚMERO")
        continue
    
    if oferta > int(mayor_oferta.iloc[0]):
        print("NÚMERO OFERTA INCORRECTO")
        continue
    else:
        break

cliente = ''
comparativo =  datos_piezas('forma_oferta', cliente , oferta)

#matriz=datos_conjuntos3.iloc[i]['MATRIZ']

extension = '.dxf'


if len(comparativo):
    



    #comparativo['TRAT'] = comparativo.apply(lambda fila: (0 if fila['VTRATMTO']<= 0 else fila['VTRATMTO'] ), axis = 1)
    comparativo= comparativo.fillna(0)
    comparativo['MATERIAL2'] = comparativo.apply(lambda fila: (10000000000 if fila['VPZ']<=0  else fila['VPZ'] ), axis = 1)
    comparativo['MATERIAL'] = comparativo.apply(lambda fila: (fila['MATERIAL2'] if fila['MATERIAL2']<=fila['PROPMAT']  else fila['PROPMAT'] ), axis = 1)
    comparativo['SUMA'] = comparativo.apply(lambda fila: (fila['MATERIAL'] + fila['VCORTE'] + fila['VTRATMTO'] + fila['PR_TRANSPORTE'] + fila['VGAS']), axis = 1)
    comparativo['SUBTOTAL_NO_VAC'] = comparativo.apply(lambda fila: (fila['QPZ'] * fila['SUMA']), axis = 1)
    comparativo['SUBTOTAL_VAC'] = comparativo.apply(lambda fila: (fila['QPZ'] * fila['VPU']), axis = 1)
    comparativo['DIFERENCIA'] = comparativo.apply(lambda fila: (fila['SUBTOTAL_NO_VAC'] - fila['SUBTOTAL_VAC']), axis = 1)
    comparativo= comparativo.drop(['MATERIAL2'], axis=1)
    
    comparativo['COEF_APROV'] = ''
    lista_coef_aprov = list()

    for i in range(len(comparativo)):
        referencia = comparativo.iloc[i]['REF_N']
        material_buscable = ['PORTE', 'TUBO', 'NULO', 'DIGIT', 'ABONO', 'NO', '~~']
        material_pieza = comparativo.iloc[i]['TIPOM']
        if material_pieza not in material_buscable:
            resultado = dato.existe_archivo(referencia, '.dxf')

            if resultado:
                ruta_archivo = 'C:\\activa\\PKS\\' + referencia + extension
                coef_aprovechamiento = plano.calcula_area(ruta_archivo)
                lista_coef_aprov.append(coef_aprovechamiento)
            else:
                lista_coef_aprov.append(0)
        else:
            lista_coef_aprov.append(0)

    comparativo= comparativo.drop(['TIPOM'], axis=1)
    comparativo['COEF_APROV'] = lista_coef_aprov
    comparativo= comparativo.round(2)
    comparativo.rename(columns={'VPZ':'MAT_MAN', 'QPZ':'CANT', 'C_VR':'VERSION'}, inplace=True)


    volcador.volcado_con_pandas(comparativo,'OFERTA',ruta2, "b" )

    wb = op.load_workbook(ruta2)
    ws = wb.active


    longitud = len(ws['A'])
    print(longitud)

    lista_columna = ['R']
    for letra in lista_columna:
        campo = letra + '2:' + letra + str(longitud)

        redFill = PatternFill(start_color='FAA506',
                       end_color='FAA506',
                       fill_type='solid')
        greenfill = PatternFill(start_color='00F000',
                       end_color='00F000',
                       fill_type='solid' )
        ws.conditional_formatting.add(campo,
            FormulaRule(formula=[(letra +'2<-1')], fill=redFill))
        ws.conditional_formatting.add(campo,
            FormulaRule(formula=[(letra + '2>=1')], fill=greenfill))

    letra = 'S'

    campo = letra + '2:' + letra + str(longitud)

    relleno_amarillo = PatternFill(start_color='CCFF66',
                       end_color='CCFF66',
                       fill_type='solid')

    ws.conditional_formatting.add(campo,
            FormulaRule(formula=[(letra +'2<0.77')], fill=relleno_amarillo))
        




    wb.save(ruta2)
    #volcador.correo(ruta2,'RESP. AUTOMATICA', 'RESUMEN.xlsx', destino)